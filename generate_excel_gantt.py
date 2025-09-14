
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
from io import BytesIO

def generate_excel_gantt(data):
    # Créer un DataFrame pandas avec les tâches
    tasks = []
    start_date = datetime.now()
    current_date = start_date

    # Phase 1: Initialisation
    tasks.append(["Initialisation du projet", current_date, current_date + timedelta(days=5), 5, 0])
    current_date += timedelta(days=5)

    # Phase 2: Conception et Design
    tasks.append(["Conception et Design", current_date, current_date + timedelta(days=10), 10, 0])
    current_date += timedelta(days=10)

    # Phase 3: Développement MVP
    tasks.append(["Développement MVP", current_date, current_date + timedelta(days=20), 20, 0])
    current_date += timedelta(days=20)

    # Phase 4: Tests et Qualité
    tasks.append(["Tests et Qualité", current_date, current_date + timedelta(days=10), 10, 0])
    current_date += timedelta(days=10)

    # Phase 5: Déploiement et Lancement
    tasks.append(["Déploiement et Lancement", current_date, current_date + timedelta(days=5), 5, 0])
    current_date += timedelta(days=5)

    # Phase 6: Suivi Post-Lancement
    tasks.append(["Suivi Post-Lancement", current_date, current_date + timedelta(days=10), 10, 0])

    df = pd.DataFrame(tasks, columns=["Tâche", "Date de début", "Date de fin", "Durée", "Complétion"])

    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan d'Action - Gantt"

    # Ajouter les en-têtes
    headers = ["Tâche", "Date de début", "Date de fin", "Durée", "Complétion"]
    ws.append(headers)

    # Ajouter les données
    for r_idx, row in enumerate(df.iterrows(), 2):
        for c_idx, value in enumerate(row[1], 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Créer le diagramme de Gantt
    chart = BarChart()
    chart.type = "bar"
    chart.y_axis.majorGridlines = None
    chart.y_axis.title = "Tâches"
    chart.x_axis.title = "Date"
    chart.x_axis.major_tick_mark = "out"
    chart.x_axis.number_format = 'd-mmm'

    data = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "Diagramme de Gantt"
    chart.legend = None

    # Masquer la première série (Date de début)
    chart.series[0].graphicalProperties.line.noFill = True
    chart.series[0].graphicalProperties.solidFill = "FFFFFF"

    ws.add_chart(chart, "G2")

    # Sauvegarder le classeur en mémoire
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream

if __name__ == "__main__":
    # Exemple de données (à remplacer par les données réelles de l'application React)
    example_data = {
        "projectName": "Plateforme E-Learning Entreprise",
        # ... (autres données)
    }
    file_stream = generate_excel_gantt(example_data)
    with open("plan_action_gantt.xlsx", "wb") as f:
        f.write(file_stream.getbuffer())


